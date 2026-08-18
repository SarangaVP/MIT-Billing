"""
One-time import: reads the 'Master sheet' tab from the existing manually
maintained Excel file and loads it into employees + mobile_numbers.

The sheet's main table (row 1 header: Mobile No | EMP No | Name | LOB |
Cadre | Credit Limit | Level | Email | ...) is the ONLY section read.

A second block below it, labeled "Additional common connection", was
originally treated as containing extra numbers not present in the main
table — but a direct check against the real file confirmed EVERY one of
its 9 real rows is already a duplicate of something in the main table
(0 genuinely unique rows). It's a redundant confirmation listing, not a
source of additional data, so it's deliberately not read at all.

Some EMP Nos in the main table are the literal text "General" — these
represent pooled "General" lines (e.g. security post phones, a shared
data bucket) not tied to one named person. Confirmed real case: 5 such
rows (Security 1/3/4, Driver Perera, Data bucket) all share this literal
EMP No — without excluding them, they'd get grouped together as one
garbled "employee" (whichever row's Name/LOB happened to be read first
stands in for all 5). Each gets its OWN synthetic EMP No (keyed by mobile
number, since these are distinct unrelated lines, not one person) and is
stored as a real employee with is_general_line=True — this used to just report
them and discard the data, which meant any bill charge on one of these
lines showed as an unattributed "Unmatched number" instead of its real
label.

Only NaN / a truly empty cell is converted to NULL. Everything else —
including literal "#N/A" text, 0s, or stray characters — is imported
exactly as it appears in the source sheet.

Usage:
    python scripts/import_master_sheet.py /path/to/Mobile_bill_July_26.xlsx
"""
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))

from app.database import SessionLocal, Base, engine  # noqa: E402
from app.models.employee import Employee  # noqa: E402
from app.models.mobile_number import MobileNumber  # noqa: E402

ADDITIONAL_BLOCK_MARKER = "Additional common connection"
UNRESOLVED_EMP_NO_PLACEHOLDERS = {"General"}
NO_NUMBER_ASSIGNED_PLACEHOLDERS = {"no", "none", "n/a", "na", "tbc", "pending", "-"}


def is_no_number_placeholder(value) -> bool:
    return isinstance(value, str) and value.strip().lower() in NO_NUMBER_ASSIGNED_PLACEHOLDERS


def split_name_and_project_label(raw_name):
    """
    Some rows encode a per-number project cost-allocation directly in the
    Name column (e.g. "Panitha Bimsara-IPTV Project") — confirmed real for
    3 employees (Panitha Bimsara, Upul Naulla, Ravindra Dharmasena): some
    of their mobile numbers are cost-allocated to a specific project,
    some aren't. Previously the employee's OWN name got silently
    overwritten by whichever row happened to be read first, since
    Employee.name is one shared field, not per-number. This keeps the
    employee's name clean and consistent, while preserving the per-number
    project label as its own field.
    """
    if raw_name is None:
        return None, None
    if not isinstance(raw_name, str):
        # A Name cell holding a stray number — treat it as a plain name
        # with no project label rather than crashing on "-" in <non-string>.
        raw_name = str(raw_name)
    if "-" not in raw_name:
        return raw_name.strip(), None
    base, _, label = raw_name.partition("-")
    return base.strip(), label.strip() or None


def clean(value):
    """Converts a truly empty cell to None, and strips invisible BOM
    characters (U+FEFF) that appear stuck in some names from copy-pasting.
    Everything else is left exactly as-is."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.replace("\ufeff", "")
        if value.strip() == "":
            return None
    return value


def to_text(value):
    if value is None:
        return None
    return str(value)


def load_main_table_rows(xlsx_path: str) -> list[dict]:
    """
    Returns a list of dicts (one per real data row), keyed by normalized
    header name — e.g. row["mobile no"], row["emp no"] — instead of
    fixed positions. Confirmed real risk this replaces: the previous
    version read row[0] through row[8] with ZERO verification that
    column 4 was actually "Cadre", column 7 actually "Email", etc. — if
    the sheet's column order ever changed, this would have silently
    corrupted data with no error at all (the same failure mode already
    confirmed and fixed for the Portal sheet elsewhere in this project).

    Also handles a CONFIRMED real typo in the source file's own header:
    the Email column is literally spelled "Emaill" (double-L), not
    "Email" — both spellings are accepted and mapped to the same field,
    so a future corrected spelling wouldn't silently break this either.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Master sheet"]

    header_row_num = None
    col_map: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if "mobile no" in values and "emp no" in values and "name" in values:
            header_row_num = row[0].row
            col_map = {v: i for i, v in enumerate(values) if v}
            break
    if header_row_num is None:
        raise ValueError("Could not find a header row containing 'Mobile No', 'EMP No', and 'Name' in the first 5 rows")

    email_idx = col_map.get("email", col_map.get("emaill"))   # confirmed real header typo: "Emaill"

    field_columns = {
        "mobile no": col_map.get("mobile no"),
        "emp no": col_map.get("emp no"),
        "name": col_map.get("name"),
        "lob": col_map.get("lob"),
        "cadre": col_map.get("cadre"),
        "credit limit": col_map.get("credit limit"),
        "level": col_map.get("level"),
        "email": email_idx,
        "resignation": col_map.get("resignation"),
    }

    all_rows = list(ws.iter_rows(min_row=header_row_num + 1, max_row=ws.max_row, values_only=True))

    marker_idx = None
    for i, row in enumerate(all_rows):
        for cell in row:
            if isinstance(cell, str) and cell.strip() == ADDITIONAL_BLOCK_MARKER:
                marker_idx = i
                break
        if marker_idx is not None:
            break

    main_rows = all_rows[:marker_idx] if marker_idx is not None else all_rows

    def row_to_dict(row: tuple) -> dict:
        return {
            field: (row[idx] if idx is not None and idx < len(row) else None)
            for field, idx in field_columns.items()
        }

    return [row_to_dict(row) for row in main_rows]


def main(xlsx_path: str):
    main_rows = load_main_table_rows(xlsx_path)

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    inserted_employees, inserted_numbers = 0, 0
    skipped_missing = 0
    skipped_duplicate_employees = 0
    duplicate_number_details = []
    unresolved_placeholder_rows = []
    no_number_employees = []

    try:
        existing_emp_nos = {e.emp_no for e in db.query(Employee.emp_no).all()}
        existing_mobile_nos = {m.mobile_no for m in db.query(MobileNumber.mobile_no).all()}

        # Column order is now resolved by header name, not position — see load_main_table_rows.
        grouped = {}
        for row in main_rows:
            raw_mobile_no = clean(row["mobile no"])
            emp_no = to_text(clean(row["emp no"]))
            raw_name = clean(row["name"])
            base_name, project_label = split_name_and_project_label(raw_name)

            if not emp_no:
                skipped_missing += 1
                continue

            if emp_no in UNRESOLVED_EMP_NO_PLACEHOLDERS:
                # A "General" line, not a real named employee — the
                # source sheet uses the literal text "General" for these
                # instead of a real EMP No. Each gets its own synthetic
                # EMP No (keyed by mobile number, since these 5 rows are
                # 5 DISTINCT unrelated lines, not one person — grouping
                # them by the shared literal "General" text was exactly
                # the original bug). Uses the SAME row-reading logic as a
                # regular employee, so real values already in the sheet
                # (credit_limit, cadre, etc.) are preserved as-is.
                mobile_no_str = to_text(raw_mobile_no)
                if not mobile_no_str:
                    skipped_missing += 1
                    continue

                synthetic_emp_no = f"GENERAL-{mobile_no_str}"
                unresolved_placeholder_rows.append((mobile_no_str, emp_no, base_name))

                if synthetic_emp_no in existing_emp_nos or mobile_no_str in existing_mobile_nos:
                    continue  # already imported on a previous run

                general_line_employee = Employee(
                    emp_no=synthetic_emp_no,
                    name=base_name,
                    lob=clean(row["lob"]),
                    cadre=clean(row["cadre"]),
                    credit_limit=clean(row["credit limit"]),
                    level=clean(row["level"]),
                    email=clean(row["email"]),
                    resignation=clean(row["resignation"]),
                    is_general_line=True,
                )
                db.add(general_line_employee)
                db.flush()
                db.add(MobileNumber(employee_id=general_line_employee.id, mobile_no=mobile_no_str, is_primary=True))
                existing_emp_nos.add(synthetic_emp_no)
                existing_mobile_nos.add(mobile_no_str)
                inserted_employees += 1
                inserted_numbers += 1
                continue

            grouped.setdefault(emp_no, {"rows": [], "first": row, "clean_name_row": None})

            # Prefer a row with NO project label for the employee's own
            # name/LOB/cadre/etc — confirmed real cases (Panitha Bimsara,
            # Upul Naulla, Ravindra Dharmasena) always have at least one
            # clean row alongside their project-labeled one(s).
            if project_label is None and grouped[emp_no]["clean_name_row"] is None:
                grouped[emp_no]["clean_name_row"] = row

            if raw_mobile_no is None:
                skipped_missing += 1
                continue

            if is_no_number_placeholder(raw_mobile_no):
                # A real employee with no company mobile number — keep the
                # employee, just don't record a mobile number for them.
                continue

            grouped[emp_no]["rows"].append((to_text(raw_mobile_no), project_label))

        for emp_no, data in grouped.items():
            if emp_no in existing_emp_nos:
                skipped_duplicate_employees += 1
                continue

            # Use the clean (no project label) row for the employee's own
            # identity fields if one exists; otherwise fall back to
            # whichever row came first, stripping its label from the name.
            source_row = data["clean_name_row"] or data["first"]
            clean_source_name, _ = split_name_and_project_label(clean(source_row["name"]))

            employee = Employee(
                emp_no=emp_no,
                name=clean_source_name,
                lob=clean(source_row["lob"]),
                cadre=clean(source_row["cadre"]),
                credit_limit=clean(source_row["credit limit"]),
                level=clean(source_row["level"]),
                email=clean(source_row["email"]),
                resignation=clean(source_row["resignation"]),
            )
            db.add(employee)
            db.flush()
            inserted_employees += 1

            if not data["rows"]:
                no_number_employees.append((emp_no, employee.name))

            # De-dupe by mobile number within this employee, preserve order.
            # If the SAME number appears more than once with different
            # project labels (shouldn't happen, but don't crash), keep the
            # first label seen.
            seen_numbers: dict[str, str | None] = {}
            for mobile_no, project_label in data["rows"]:
                if mobile_no not in seen_numbers:
                    seen_numbers[mobile_no] = project_label

            for i, (mobile_no, project_label) in enumerate(seen_numbers.items()):
                if mobile_no in existing_mobile_nos:
                    duplicate_number_details.append((emp_no, mobile_no, "already used elsewhere"))
                    continue
                db.add(MobileNumber(
                    employee_id=employee.id, mobile_no=mobile_no, is_primary=(i == 0),
                    project_label=project_label,
                ))
                existing_mobile_nos.add(mobile_no)
                inserted_numbers += 1

        db.commit()
    finally:
        db.close()

    print(f"Imported {inserted_employees} employees with {inserted_numbers} mobile numbers total.")
    print(f"Skipped {skipped_missing} rows with missing EMP No or mobile no.")
    print(f"Skipped {skipped_duplicate_employees} EMP Nos that already existed in the database.")

    if duplicate_number_details:
        print(f"\n{len(duplicate_number_details)} mobile number(s) could not be attached (already claimed elsewhere):")
        for emp_no, mobile_no, reason in duplicate_number_details:
            print(f"  - {mobile_no} for EMP No {emp_no} ({reason})")

    if unresolved_placeholder_rows:
        print(f"\n{len(unresolved_placeholder_rows)} 'General' line(s) imported as their own employee record (EMP No 'General' in the source sheet):")
        for mobile_no, emp_no, label in unresolved_placeholder_rows:
            print(f"  - {mobile_no} -> {label} (synthetic EMP No: GENERAL-{mobile_no})")

    if no_number_employees:
        print(f"\n{len(no_number_employees)} employee(s) imported with NO mobile number (source sheet had a placeholder like 'No'):")
        for emp_no, name in no_number_employees:
            print(f"  - {name} (EMP No: {emp_no})")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python scripts/import_master_sheet.py /path/to/file.xlsx")
        sys.exit(1)
    main(sys.argv[1])