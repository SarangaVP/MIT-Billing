"""
One-time import: reads the 'Master sheet' from the Dialog Data Bucket
Excel export and seeds dialog_data_employees + dialog_data_connections.

CONFIRMED two real file formats seen in practice, same pattern as
Mobitel's Summary sheet:
  - Newer format: 'Master sheet' has a 5th column, "LOB", with the
    numeric team code directly on each row — no cross-referencing
    needed at all. Verified the values agree exactly with the older
    format's separate 'LOB' sheet (e.g. Mahesh Wijenayaka: 76 both ways).
  - Older format: 'Master sheet' has only 4 columns (Connection No,
    EMP No, Employee, Team) — lob_code must be cross-referenced from the
    separate 'LOB' sheet (an org-chart/HR export), joined by EMP No.

Reads columns by HEADER NAME, not fixed position, so this correctly
detects which format a given file is — matches the same defensive
pattern used for Mobitel's Summary sheet after that file's columns
shifted between months.

CONFIRMED the separate 'LOB' sheet (used only for the older format) has
TWO columns both literally named "LOB" — the first has 33 rows with a
literal "#N/A" error string, the second has zero errors. Uses the
second one. Coverage there is ~95% (467/489 employees) — people with no
matching row simply get lob_code=None.

CONFIRMED against the real file: an employee can hold MORE THAN ONE
connection (e.g. "Indusarani Silva" appears twice with 2 different
connection numbers, each billed separately at the full rate) — so rows
are grouped by EMP No first, then each row's connection number is added
under that one employee, mirroring Dialog Mobile's Employee/MobileNumber
structure, NOT Mobitel's simpler 1:1 model.

Usage:
    python scripts/import_dialog_data_master.py /path/to/Dialog_data_Jul26.xlsx
"""
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))

from app.database import SessionLocal, Base, engine  # noqa: E402
from app.models.dialog_data_employee import DialogDataEmployee  # noqa: E402
from app.models.dialog_data_connection import DialogDataConnection  # noqa: E402


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = re.sub(r"\s+", " ", value.replace("\ufeff", "")).strip()
        return cleaned or None
    return value


def to_str(value) -> str | None:
    cleaned = clean(value)
    if cleaned is None:
        return None
    text = str(int(cleaned)) if isinstance(cleaned, float) else str(cleaned)
    # "#N/A" is a genuine unresolved-lookup marker (confirmed real). "0" is
    # also confirmed real as a placeholder for an unresolved LOB code in
    # at least one file — not a genuine business code — so both are
    # treated as "no value" rather than imported as if they were real.
    if text in ("#N/A", "0"):
        return None
    return text


def _build_column_map(header_row) -> dict[str, int]:
    """Maps normalized header text -> column index, so we never rely on fixed positions."""
    col_map = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        col_map[key] = idx
    return col_map


def _load_lob_codes_from_separate_sheet(wb) -> dict[str, str]:
    """
    Fallback for the OLDER format — cross-references the separate 'LOB'
    sheet by EMP No. Confirmed this sheet's real header uses 'EMP#', not
    'EMP No', AND has TWO columns both literally named 'LOB' (the first
    has real #N/A errors, the second is reliable). Since dict
    comprehension keeps the LAST occurrence for a repeated key, mapping
    by header name lands on the second (reliable) "LOB" column
    automatically, without needing a hardcoded position — documented
    explicitly here so it's not a silent coincidence.
    """
    if "LOB" not in wb.sheetnames:
        return {}
    ws = wb["LOB"]

    header_row_num = None
    col_map: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if "emp#" in values and "name" in values:
            header_row_num = row[0].row
            col_map = {v: i for i, v in enumerate(values) if v}
            break
    if header_row_num is None:
        return {}

    emp_no_idx = col_map.get("emp#")
    lob_idx = col_map.get("lob")   # last occurrence wins -> the reliable second "LOB" column
    if emp_no_idx is None or lob_idx is None:
        return {}

    codes: dict[str, str] = {}
    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=1000, values_only=True):
        emp_no = to_str(row[emp_no_idx]) if emp_no_idx < len(row) else None
        lob_code = to_str(row[lob_idx]) if lob_idx < len(row) else None
        if emp_no and lob_code:
            codes[emp_no] = lob_code
    return codes


def _find_header_row(ws) -> tuple[int, list]:
    """Searches the first few rows for the real header, rather than
    assuming it's always exactly row 1 — same defensive pattern applied
    across every sheet in this project after confirming real title-row
    shifts happen in this workbook family."""
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if "connection no" in values and "emp no" in values and "employee" in values:
            return row[0].row, [c.value for c in row]
    raise ValueError("Could not find a header row containing 'Connection No', 'EMP No', and 'Employee' in the first 5 rows")


def main(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Master sheet"]

    header_row_num, header_row = _find_header_row(ws)
    col_map = _build_column_map(header_row)

    connection_idx = col_map.get("connection no")
    emp_no_idx = col_map.get("emp no")
    name_idx = col_map.get("employee")
    team_idx = col_map.get("team")
    lob_idx = col_map.get("lob")   # present only in the newer format

    has_lob_in_master = lob_idx is not None
    lob_codes_from_separate_sheet = {} if has_lob_in_master else _load_lob_codes_from_separate_sheet(wb)

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    employees_inserted, connections_inserted, lob_codes_backfilled = 0, 0, 0
    skipped_missing, skipped_duplicate_connection = 0, 0

    try:
        employee_by_emp_no: dict[str, DialogDataEmployee] = {
            e.emp_no: e for e in db.query(DialogDataEmployee).all()
        }
        existing_connection_nos = {
            c.connection_no for c in db.query(DialogDataConnection).all()
        }

        for row in ws.iter_rows(min_row=header_row_num + 1, max_row=1000, values_only=True):
            connection_no = row[connection_idx] if connection_idx is not None and connection_idx < len(row) else None
            emp_no = row[emp_no_idx] if emp_no_idx is not None and emp_no_idx < len(row) else None
            name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
            team = row[team_idx] if team_idx is not None and team_idx < len(row) else None

            if connection_no is None and emp_no is None and name is None:
                continue

            connection_no_clean, emp_no_clean, name_clean, team_clean = (
                clean(connection_no), clean(emp_no), clean(name), clean(team)
            )

            # Some rows have the literal text "No" typed into Connection No,
            # meaning "no connection assigned" rather than a real number —
            # confirmed by a real duplicate-key crash when two different
            # employees both had the literal string "No" here. Treat it as
            # missing, same as blank/0/#N/A elsewhere in this script.
            if isinstance(connection_no_clean, str) and connection_no_clean.strip().lower() == "no":
                connection_no_clean = None

            if not connection_no_clean or not emp_no_clean or not name_clean:
                skipped_missing += 1
                continue

            connection_no_str = (
                str(int(connection_no_clean)) if isinstance(connection_no_clean, float) else str(connection_no_clean)
            )
            emp_no_str = str(emp_no_clean)

            if has_lob_in_master:
                lob_code = to_str(row[lob_idx]) if lob_idx < len(row) else None
            else:
                lob_code = lob_codes_from_separate_sheet.get(emp_no_str)

            if connection_no_str in existing_connection_nos:
                skipped_duplicate_connection += 1
                continue

            employee = employee_by_emp_no.get(emp_no_str)
            if employee is None:
                employee = DialogDataEmployee(
                    emp_no=emp_no_str, name=name_clean, team=team_clean,
                    lob_code=lob_code,
                )
                db.add(employee)
                db.flush()
                employee_by_emp_no[emp_no_str] = employee
                employees_inserted += 1
                if employee.lob_code:
                    lob_codes_backfilled += 1
            elif employee.lob_code is None and lob_code:
                # Backfill lob_code for an employee already seeded before
                # this field existed, or before the file had this column.
                employee.lob_code = lob_code
                lob_codes_backfilled += 1

            db.add(DialogDataConnection(employee_id=employee.id, connection_no=connection_no_str))
            existing_connection_nos.add(connection_no_str)
            connections_inserted += 1

        db.commit()
    finally:
        db.close()

    print(f"Imported {employees_inserted} Dialog Data Bucket employees.")
    print(f"Imported {connections_inserted} connections (some employees have more than one).")
    print(f"Backfilled lob_code for {lob_codes_backfilled} employees.")
    print(f"Skipped {skipped_missing} rows with missing connection no/EMP No/name.")
    print(f"Skipped {skipped_duplicate_connection} duplicate connection numbers.")
    print(f"LOB source: {'directly in Master sheet' if has_lob_in_master else 'separate LOB sheet (older format)'}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python scripts/import_dialog_data_master.py /path/to/file.xlsx")
        sys.exit(1)
    main(sys.argv[1])