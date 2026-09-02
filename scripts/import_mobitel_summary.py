"""
One-time import (and re-run-safe refresh): reads the 'Master sheet' from
the Mobitel data bucket Excel export and seeds/updates mobitel_employees
+ mobitel_connections.

IMPORTANT: reads columns by HEADER NAME, not fixed position. Only
"Number", "EMP No", and "Name" are hard requirements. Neither "Team"
(team name) nor "LOB" (numeric code) is required — if "Team" is missing,
team simply stays None; if "LOB" is missing, the code is looked up from
the separate "LOB" sheet by EMP No instead (same fallback pattern as
Dialog Data Bucket). The older Mobitel format, where a single "LOB"
column held the team name with no numeric code at all, is not in active
use, so no special handling is needed for that ambiguity.

Rows where EMP No is 'NA' and Name is 'Pool' are unassigned SIMs — real
lines held by no one, not employees. These ARE imported (for visibility),
as a synthetic employee (emp_no "POOL-<mobile_no>", is_pool=True) with
one connection.

Rows are grouped by EMP No — an employee can hold MORE THAN ONE
connection. No such case exists in the real June/July data (checked
directly), but the schema and this script now support it correctly if
it ever does: previously, a duplicate EMP No with a different mobile
number was silently dropped with zero warning, which was a real,
unflagged gap.

Re-running this script with a newer file UPDATES existing employees'
team/lob_code (e.g. to backfill lob_code for people seeded before it
existed) and ADDS any new connection numbers found for them — it does
not just skip duplicates blindly.

Usage:
    python scripts/import_mobitel_summary.py /path/to/Mobitel_Jul26.xlsx
"""
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))

from app.database import SessionLocal, Base, engine  # noqa: E402
from app.models.mobitel_employee import MobitelEmployee  # noqa: E402
from app.models.mobitel_connection import MobitelConnection  # noqa: E402

POOL_MARKERS = {"na", "pool"}


def is_pool_row(emp_no, name) -> bool:
    return (str(emp_no).strip().lower() in POOL_MARKERS) or (str(name).strip().lower() == "pool")


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = re.sub(r"\s+", " ", value.replace("\ufeff", "")).strip()
        return cleaned or None
    return value


def to_str(value) -> str | None:
    cleaned = clean(value)
    if cleaned is None:
        return None
    return str(int(cleaned)) if isinstance(cleaned, float) else str(cleaned)


def _build_column_map(header_row) -> dict[str, int]:
    """Maps normalized header text -> column index, so we never rely on fixed positions."""
    col_map = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        col_map[key] = idx
    return col_map


def _find_header_row(ws) -> tuple[int, list]:
    """Searches the first few rows for the real header, rather than
    assuming it's always exactly row 1 — same defensive pattern as the
    Portal sheet fix (that sheet, in this same workbook, once had a
    title row above its header, which a fixed-row assumption would have
    silently misread)."""
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if "number" in values and "emp no" in values and "name" in values:
            return row[0].row, [c.value for c in row]
    raise ValueError("Could not find a header row containing 'Number', 'EMP No', and 'Name' in the first 5 rows")


def _load_lob_codes_from_separate_sheet(wb) -> dict[str, str]:
    """
    Fallback for when the Master sheet has NO "LOB" (numeric code) column
    at all — cross-references the separate 'LOB' sheet by EMP No instead,
    same pattern as Dialog Data Bucket's own fallback. Confirmed real:
    this sheet's header uses 'EMP#' (not 'EMP No'), and its code column
    is literally named 'LOB' (not 'LOB Code') — matched by header name,
    not position.

    NOTE: the older Mobitel format, where a column literally called
    "LOB" held the team NAME instead of a numeric code, is not in active
    use — so there's no need to guard against that ambiguity here.
    """
    if "LOB" not in wb.sheetnames:
        return {}
    ws = wb["LOB"]

    header_row_num = None
    col_map: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if "emp#" in values and "name" in values:
            header_row_num = row[0].row
            col_map = {v: i for i, v in enumerate(values) if v}
            break
    if header_row_num is None:
        return {}

    emp_idx = col_map.get("emp#")
    lob_idx = col_map.get("lob")
    if emp_idx is None or lob_idx is None:
        return {}

    codes: dict[str, str] = {}
    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=ws.max_row, values_only=True):
        emp = row[emp_idx] if emp_idx < len(row) else None
        code = row[lob_idx] if lob_idx < len(row) else None
        if emp is not None and code is not None:
            emp_str = str(int(emp)) if isinstance(emp, float) else str(emp)
            codes[emp_str] = str(int(code)) if isinstance(code, float) else str(code)
    return codes


def main(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Master sheet"]

    header_row_num, header_row = _find_header_row(ws)
    col_map = _build_column_map(header_row)

    mobile_idx = col_map.get("number")
    emp_no_idx = col_map.get("emp no")
    name_idx = col_map.get("name")
    team_idx = col_map.get("team")
    lob_idx = col_map.get("lob")
    has_lob_in_master = lob_idx is not None

    if mobile_idx is None or emp_no_idx is None or name_idx is None:
        print("ERROR: could not find 'Number'/'EMP No'/'Name' columns in the header row — check the file format.")
        print(f"Header found: {header_row}")
        sys.exit(1)

    # Fallback: the Master sheet may not have its own numeric "LOB" code
    # column at all — cross-reference the separate "LOB" sheet by EMP No
    # instead, same pattern as Dialog Data Bucket's own fallback. Neither
    # "Team" nor "LOB" is a hard requirement — both simply default to
    # None/fallback if absent, matching Dialog Data Bucket's own leniency.
    lob_codes_from_separate_sheet = {} if has_lob_in_master else _load_lob_codes_from_separate_sheet(wb)

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    inserted, inserted_pool, connections_added, updated = 0, 0, 0, 0
    skipped_missing, conflicts = 0, 0
    conflict_details = []

    try:
        employee_by_emp_no: dict[str, MobitelEmployee] = {
            e.emp_no: e for e in db.query(MobitelEmployee).all()
        }
        employee_by_mobile: dict[str, MobitelEmployee] = {}
        for conn in db.query(MobitelConnection).filter(MobitelConnection.is_deleted.is_(False)).all():
            employee_by_mobile[conn.mobile_no] = conn.employee

        for row in ws.iter_rows(min_row=header_row_num + 1, max_row=1000, values_only=True):
            mobile_no = row[mobile_idx] if mobile_idx < len(row) else None
            emp_no = row[emp_no_idx] if emp_no_idx < len(row) else None
            name = row[name_idx] if name_idx < len(row) else None

            if mobile_no is None and emp_no is None and name is None:
                continue  # blank filler row past the real data

            mobile_no_str = to_str(mobile_no)

            if is_pool_row(emp_no, name):
                if not mobile_no_str:
                    skipped_missing += 1
                    continue
                synthetic_emp_no = f"POOL-{mobile_no_str}"
                if synthetic_emp_no in employee_by_emp_no:
                    continue  # already imported as a pool row

                conflicting_owner = employee_by_mobile.get(mobile_no_str)
                if conflicting_owner is not None:
                    # This number was a real employee's line before, now
                    # shows as unassigned "Pool" in this file — a genuine
                    # status change, not something to silently overwrite.
                    conflicts += 1
                    conflict_details.append(
                        f"  {mobile_no_str}: file shows this as unassigned 'Pool', but it's currently "
                        f"assigned to EMP {conflicting_owner.emp_no} ({conflicting_owner.name}) in our records"
                    )
                    continue

                employee = MobitelEmployee(emp_no=synthetic_emp_no, name="Pool", is_pool=True)
                db.add(employee)
                db.flush()
                db.add(MobitelConnection(employee_id=employee.id, mobile_no=mobile_no_str))
                employee_by_emp_no[synthetic_emp_no] = employee
                employee_by_mobile[mobile_no_str] = employee
                inserted_pool += 1
                continue

            emp_no_str, name_clean = to_str(emp_no), clean(name)
            if not mobile_no_str or not emp_no_str or not name_clean:
                skipped_missing += 1
                continue

            team_name = clean(row[team_idx]) if team_idx is not None and team_idx < len(row) else None
            lob_code = to_str(row[lob_idx]) if has_lob_in_master and lob_idx < len(row) else lob_codes_from_separate_sheet.get(emp_no_str)

            existing = employee_by_emp_no.get(emp_no_str)
            if existing is None:
                # This EMP No is new to us — but check whether this mobile
                # number is already claimed by a DIFFERENT existing
                # employee (confirmed real scenario: a number can be
                # reassigned to someone else between months). Don't guess
                # what to do — report it clearly instead of crashing or
                # silently overwriting someone else's record.
                conflicting_owner = employee_by_mobile.get(mobile_no_str)
                if conflicting_owner is not None:
                    conflicts += 1
                    conflict_details.append(
                        f"  {mobile_no_str}: file says EMP {emp_no_str} ({name_clean}), "
                        f"but already assigned to EMP {conflicting_owner.emp_no} ({conflicting_owner.name})"
                    )
                    continue

                employee = MobitelEmployee(emp_no=emp_no_str, name=name_clean, lob=team_name, lob_code=lob_code)
                db.add(employee)
                db.flush()
                db.add(MobitelConnection(employee_id=employee.id, mobile_no=mobile_no_str))
                employee_by_emp_no[emp_no_str] = employee
                employee_by_mobile[mobile_no_str] = employee
                inserted += 1
            else:
                # Refresh team/lob_code on re-import (e.g. to backfill
                # lob_code for people originally seeded from an older
                # file that didn't have this column).
                changed = False
                if team_name and existing.lob != team_name:
                    existing.lob = team_name
                    changed = True
                if lob_code and existing.lob_code != lob_code:
                    existing.lob_code = lob_code
                    changed = True
                if changed:
                    updated += 1

                # This is the actual fix: if this employee doesn't yet
                # have THIS mobile number as one of their connections,
                # add it as a NEW connection instead of silently dropping
                # it (confirmed the previous version of this script did
                # exactly that — tested with a synthetic multi-mobile
                # case and it lost the second number with zero warning).
                if mobile_no_str not in employee_by_mobile:
                    db.add(MobitelConnection(employee_id=existing.id, mobile_no=mobile_no_str))
                    employee_by_mobile[mobile_no_str] = existing
                    connections_added += 1
                elif employee_by_mobile[mobile_no_str].emp_no != emp_no_str:
                    conflicts += 1
                    owner = employee_by_mobile[mobile_no_str]
                    conflict_details.append(
                        f"  {mobile_no_str}: file says EMP {emp_no_str} ({name_clean}), "
                        f"but already assigned to EMP {owner.emp_no} ({owner.name})"
                    )

        db.commit()
    finally:
        db.close()

    print(f"Imported {inserted} new Mobitel employees.")
    print(f"Imported {inserted_pool} new unassigned 'Pool' rows.")
    print(f"Added {connections_added} additional connections for existing employees (multi-number cases).")
    print(f"Updated {updated} existing employees (team/lob_code refreshed).")
    print(f"Skipped {skipped_missing} rows with missing EMP No/name/mobile no.")
    if conflicts:
        print(f"CONFLICTS: {conflicts} mobile number(s) claimed by a different EMP No than currently on file:")
        for line in conflict_details:
            print(line)
        print("These were NOT imported — resolve manually (likely a genuine number reassignment).")
    print(f"LOB code source: {'directly in Master sheet' if has_lob_in_master else ('separate LOB sheet' if lob_codes_from_separate_sheet else 'unavailable')}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python scripts/import_mobitel_summary.py /path/to/file.xlsx")
        sys.exit(1)
    main(sys.argv[1])