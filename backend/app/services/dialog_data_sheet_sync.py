"""
Syncs Dialog Data Bucket employees + connections from an uploaded Master
sheet, treating it as the full, authoritative current roster — same
"update in place, never delete" approach as employee_sheet_sync.py and
mobitel_sheet_sync.py (see either for why a real delete-and-replace
isn't safe: connections are referenced by past bills via a real foreign
key, and connection_no/EMP No have unique constraints that block reusing
an identifier even after a soft delete).

Shares its row-parsing logic (header-name column detection for the
newer/older LOB column formats, "0"/"#N/A" placeholder filtering) with
import_dialog_data_master.py.
"""
import openpyxl
from sqlalchemy.orm import Session

from app.models.dialog_data_employee import DialogDataEmployee
from app.models.dialog_data_connection import DialogDataConnection, DialogDataConnectionStatus


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.replace("\ufeff", "").strip()
        return cleaned or None
    return value


def to_str(value) -> str | None:
    cleaned = clean(value)
    if cleaned is None:
        return None
    text = str(int(cleaned)) if isinstance(cleaned, float) else str(cleaned)
    if text in ("#N/A", "0"):
        return None
    return text


def _build_column_map(header_row) -> dict[str, int]:
    col_map = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        col_map[str(cell).strip().lower()] = idx
    return col_map


def _load_lob_codes_from_separate_sheet(wb) -> dict[str, str]:
    """
    Fallback for the OLDER format — cross-references the separate 'LOB'
    sheet by EMP No. Confirmed this sheet's real header uses 'EMP#', not
    'EMP No', AND has TWO columns both literally named 'LOB' (the first
    has real #N/A errors, the second is reliable — same finding as the
    original investigation). Since dict comprehension keeps the LAST
    occurrence for a repeated key, mapping by header name lands on the
    second (reliable) "LOB" column automatically, without needing a
    hardcoded position — but this is documented explicitly here so it's
    not a silent coincidence.
    """
    if "LOB" not in wb.sheetnames:
        return {}
    ws = wb["LOB"]

    header_row_num = None
    col_map: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if "emp#" in values and "name" in values:
            header_row_num = row[0].row
            col_map = {v: i for i, v in enumerate(values) if v}
            break
    if header_row_num is None:
        return {}

    emp_no_idx = col_map.get("emp#")
    lob_idx = col_map.get("lob")   # last occurrence wins -> the reliable second "LOB" column
    if emp_no_idx is None or lob_idx is None:
        return {}

    codes: dict[str, str] = {}
    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=1000, values_only=True):
        emp_no = to_str(row[emp_no_idx]) if emp_no_idx < len(row) else None
        lob_code = to_str(row[lob_idx]) if lob_idx < len(row) else None
        if emp_no and lob_code:
            codes[emp_no] = lob_code
    return codes


def _find_header_row(ws) -> tuple[int, list]:
    """Searches the first few rows for the real header, rather than
    assuming it's always exactly row 1 — same defensive pattern applied
    to Mobitel's Master sheet and the Portal sheet after confirming a
    real title-row shift can happen in this workbook family."""
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if "connection no" in values and "emp no" in values and "employee" in values:
            return row[0].row, [c.value for c in row]
    raise ValueError("Could not find a header row containing 'Connection No', 'EMP No', and 'Employee' in the first 5 rows")


def sync_dialog_data_sheet(db: Session, xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Master sheet"]

    header_row_num, header_row = _find_header_row(ws)
    col_map = _build_column_map(header_row)

    connection_idx = col_map.get("connection no")
    emp_no_idx = col_map.get("emp no")
    name_idx = col_map.get("employee")
    team_idx = col_map.get("team")
    lob_idx = col_map.get("lob")

    has_lob_in_master = lob_idx is not None
    lob_codes_from_separate_sheet = {} if has_lob_in_master else _load_lob_codes_from_separate_sheet(wb)

    employee_by_emp_no: dict[str, DialogDataEmployee] = {e.emp_no: e for e in db.query(DialogDataEmployee).all()}
    connection_by_no: dict[str, DialogDataConnection] = {c.connection_no: c for c in db.query(DialogDataConnection).all()}
    # Tracks emp_no by employee id so a transfer can be reported as
    # "moved from EMP X to EMP Y" — kept in sync whenever an employee is
    # looked up or newly created below.
    emp_no_by_employee_id: dict = {e.id: e.emp_no for e in employee_by_emp_no.values()}

    grouped: dict[str, dict] = {}
    skipped_missing = 0

    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=1000, values_only=True):
        connection_no = row[connection_idx] if connection_idx is not None and connection_idx < len(row) else None
        emp_no = row[emp_no_idx] if emp_no_idx is not None and emp_no_idx < len(row) else None
        name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
        team = row[team_idx] if team_idx is not None and team_idx < len(row) else None

        if connection_no is None and emp_no is None and name is None:
            continue

        connection_no_clean, emp_no_clean, name_clean, team_clean = clean(connection_no), clean(emp_no), clean(name), clean(team)

        # Some rows in the real sheet have the literal text "No" typed into
        # the Connection No column — meaning "this employee has no
        # connection assigned" rather than an actual connection number.
        # Confirmed by a real crash: two different employees both had the
        # literal string "No" here, and since it isn't empty it was being
        # treated as if it were one real, shared connection identifier,
        # violating the unique constraint on the second insert. Treat it
        # as missing, same as blank/0/#N/A are treated elsewhere in this file.
        if isinstance(connection_no_clean, str) and connection_no_clean.strip().lower() == "no":
            connection_no_clean = None

        if not connection_no_clean or not emp_no_clean or not name_clean:
            skipped_missing += 1
            continue

        connection_no_str = str(int(connection_no_clean)) if isinstance(connection_no_clean, float) else str(connection_no_clean)
        emp_no_str = str(emp_no_clean)

        lob_code = to_str(row[lob_idx]) if has_lob_in_master and lob_idx < len(row) else lob_codes_from_separate_sheet.get(emp_no_str)

        grouped.setdefault(emp_no_str, {"name": name_clean, "team": team_clean, "lob_code": lob_code, "connections": []})
        grouped[emp_no_str]["connections"].append(connection_no_str)

    emp_nos_in_upload = set(grouped.keys())
    inserted, updated, revived, retired_employees = 0, 0, 0, 0
    connections_added, connections_retired, connections_reactivated, connections_transferred = 0, 0, 0, 0
    transfers: list[str] = []

    for emp_no, data in grouped.items():
        existing = employee_by_emp_no.get(emp_no)
        if existing is None:
            employee = DialogDataEmployee(emp_no=emp_no, name=data["name"], team=data["team"], lob_code=data["lob_code"])
            db.add(employee)
            db.flush()
            employee_by_emp_no[emp_no] = employee
            emp_no_by_employee_id[employee.id] = emp_no
            inserted += 1
        else:
            employee = existing
            was_deleted = employee.is_deleted
            employee.name = data["name"]
            employee.team = data["team"]
            if data["lob_code"]:
                employee.lob_code = data["lob_code"]
            employee.is_deleted = False
            if was_deleted:
                revived += 1
            else:
                updated += 1

        current_connections = {c.connection_no: c for c in db.query(DialogDataConnection).filter(DialogDataConnection.employee_id == employee.id).all()}
        new_connections = set(data["connections"])

        for connection_no in new_connections:
            conn = connection_by_no.get(connection_no)
            if conn is None:
                new_conn = DialogDataConnection(employee_id=employee.id, connection_no=connection_no, status=DialogDataConnectionStatus.active)
                db.add(new_conn)
                # CRITICAL: register the brand-new connection back into the
                # cache immediately. Without this, if the SAME connection
                # number appears again later in this same upload — under a
                # different EMP No, a genuine duplicate/typo in the source
                # sheet — this lookup would still return None and attempt
                # a second INSERT of the same connection_no, crashing on
                # the unique constraint instead of correctly falling into
                # the transfer branch below.
                connection_by_no[connection_no] = new_conn
                connections_added += 1
            elif conn.employee_id != employee.id:
                # The sheet says this connection now belongs to a
                # different employee than our records show — auto-transfer
                # it rather than just flagging a conflict. This is safe to
                # do automatically because past bills read a FROZEN
                # snapshot of who they were billed to at import time (see
                # DialogDataBillLineItem), not this live employee_id — so
                # reassigning it here only affects bills imported from now
                # on, never anything already imported.
                previous_emp_no = emp_no_by_employee_id.get(conn.employee_id, "unknown")
                conn.employee_id = employee.id
                conn.status = DialogDataConnectionStatus.active
                transfers.append(f"{connection_no}: transferred from EMP {previous_emp_no} to EMP {emp_no}")
                connections_transferred += 1
            elif conn.status != DialogDataConnectionStatus.active:
                conn.status = DialogDataConnectionStatus.active
                connections_reactivated += 1

        for connection_no, conn in current_connections.items():
            if connection_no not in new_connections and conn.status == DialogDataConnectionStatus.active:
                conn.status = DialogDataConnectionStatus.inactive
                connections_retired += 1

    for emp_no, employee in employee_by_emp_no.items():
        if emp_no not in emp_nos_in_upload and not employee.is_deleted:
            employee.is_deleted = True
            retired_employees += 1

    db.commit()

    return {
        "inserted_employees": inserted,
        "updated_employees": updated,
        "revived_employees": revived,
        "retired_employees": retired_employees,
        "connections_added": connections_added,
        "connections_retired": connections_retired,
        "connections_reactivated": connections_reactivated,
        "connections_transferred": connections_transferred,
        "skipped_missing_rows": skipped_missing,
        "transfers": transfers,
        "lob_source": "directly in Master sheet" if has_lob_in_master else "separate LOB sheet (older format)",
    }