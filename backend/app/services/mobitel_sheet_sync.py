"""
Syncs Mobitel employees + connections from an uploaded Master sheet,
treating it as the full, authoritative current roster — same "update in
place, never delete" approach as employee_sheet_sync.py (see that
module's docstring for why a real delete-and-replace isn't safe here:
connections are referenced by past bills via a real foreign key, and
mobile_no/emp_no have unique constraints that block reusing an
identifier even after a soft delete).

Only "Number", "EMP No", and "Name" are hard requirements. Neither
"Team" (team name) nor "LOB" (numeric code) is required — if "Team" is
missing, team simply stays None; if "LOB" is missing, the code is looked
up from the separate "LOB" sheet by EMP No instead (same fallback
pattern as Dialog Data Bucket). The older Mobitel format, where a single
"LOB" column held the team name with no numeric code at all, is not in
active use, so no special handling is needed for that ambiguity.

Shares its row-parsing logic (header-name column detection, Pool row
handling) with import_mobitel_summary.py.
"""
import re

import openpyxl
from sqlalchemy.orm import Session

from app.models.mobitel_employee import MobitelEmployee
from app.models.mobitel_connection import MobitelConnection, MobitelConnectionStatus

POOL_MARKERS = {"na", "pool"}


def is_pool_row(emp_no, name) -> bool:
    return (str(emp_no).strip().lower() in POOL_MARKERS) or (str(name).strip().lower() == "pool")


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        # Trims the edges AND collapses any run of internal whitespace
        # down to a single space (not just .strip()) — confirmed real for
        # Dialog Mobile's Cadre field (e.g. "Fixed Term " or "Fixed  Term"
        # instead of "Fixed Term"), which silently broke an exact-match
        # comparison elsewhere. Applying the same normalization here
        # protects this module's own Team Cost grouping (which keys
        # directly off the raw team string) from the identical failure
        # mode.
        cleaned = re.sub(r"\s+", " ", value.replace("\ufeff", "")).strip()
        return cleaned or None
    return value


def to_str(value) -> str | None:
    cleaned = clean(value)
    if cleaned is None:
        return None
    return str(int(cleaned)) if isinstance(cleaned, float) else str(cleaned)


def split_name_and_project_label(raw_name):
    """
    Some rows encode a per-connection project allocation directly in the
    Name column (e.g. "SLA-IPTV Project"), same convention as Dialog
    Mobile's project_label handling. Keeps the employee's own name clean
    and consistent, while preserving the per-connection project label as
    its own field on MobitelConnection.
    """
    if raw_name is None:
        return None, None
    if not isinstance(raw_name, str):
        # A Name cell holding a stray number (e.g. an accidental numeric
        # value or a formula artifact) — treat it as a plain name with no
        # project label rather than crashing on "-" in <non-string>.
        raw_name = str(raw_name)
    if "-" not in raw_name:
        return raw_name.strip(), None
    base, _, label = raw_name.partition("-")
    return base.strip(), label.strip() or None


def _build_column_map(header_row) -> dict[str, int]:
    col_map = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        col_map[str(cell).strip().lower()] = idx
    return col_map


def _find_header_row(ws) -> tuple[int, list]:
    """Searches the first few rows for the real header, rather than
    assuming it's always exactly row 1 — confirmed real risk: this same
    workbook's Portal sheet once had a title row above its header,
    which would have silently misread data under a fixed-row assumption."""
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if "number" in values and "emp no" in values and "name" in values:
            return row[0].row, [c.value for c in row]
    raise ValueError("Could not find a header row containing 'Number', 'EMP No', and 'Name' in the first 5 rows")


def _load_lob_codes_from_separate_sheet(wb) -> dict[str, str]:
    """
    Fallback for when the Master sheet has NO "LOB" (numeric code) column
    at all — cross-references the separate 'LOB' sheet by EMP No instead,
    same pattern as Dialog Data Bucket's own fallback. Confirmed real:
    this sheet's header uses 'EMP#' (not 'EMP No'), and its code column
    is literally named 'LOB' (not 'LOB Code') — matched by header name,
    not position. Confirmed reliable specifically for Mobitel: a direct
    check against a real file found this sheet's codes agree with the
    Master sheet's own embedded codes 120/120 times, zero disagreements.

    NOTE: the older Mobitel format, where a column literally called
    "LOB" held the team NAME instead of a numeric code, is not in active
    use — so there's no need to guard against that ambiguity here.
    """
    if "LOB" not in wb.sheetnames:
        return {}
    ws = wb["LOB"]

    header_row_num = None
    col_map: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if "emp#" in values and "name" in values:
            header_row_num = row[0].row
            col_map = {v: i for i, v in enumerate(values) if v}
            break
    if header_row_num is None:
        return {}

    emp_idx = col_map.get("emp#")
    lob_idx = col_map.get("lob")
    if emp_idx is None or lob_idx is None:
        return {}

    codes: dict[str, str] = {}
    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=ws.max_row, values_only=True):
        emp = row[emp_idx] if emp_idx < len(row) else None
        code = row[lob_idx] if lob_idx < len(row) else None
        if emp is not None and code is not None:
            emp_str = str(int(emp)) if isinstance(emp, float) else str(emp)
            codes[emp_str] = str(int(code)) if isinstance(code, float) else str(code)
    return codes


def sync_mobitel_sheet(db: Session, xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Master sheet"]

    header_row_num, header_row = _find_header_row(ws)
    col_map = _build_column_map(header_row)

    mobile_idx = col_map.get("number")
    emp_no_idx = col_map.get("emp no")
    name_idx = col_map.get("name")
    team_idx = col_map.get("team")
    lob_idx = col_map.get("lob")
    has_lob_in_master = lob_idx is not None

    if mobile_idx is None or emp_no_idx is None or name_idx is None:
        raise ValueError("Could not find 'Number'/'EMP No'/'Name' columns in the header row.")

    # Fallback: the Master sheet may not have its own numeric "LOB" code
    # column at all — cross-reference the separate "LOB" sheet by EMP No
    # instead, same pattern as Dialog Data Bucket's own fallback. Only
    # loaded when actually needed; the Master sheet's own embedded code
    # always wins outright when present. Neither "Team" nor "LOB" is a
    # hard requirement — both simply default to None/fallback if absent,
    # matching Dialog Data Bucket's own leniency. The older Mobitel
    # format (where "LOB" held the team name with no numeric code) is not
    # in active use, so the historical ambiguity that once justified
    # requiring "Team" as a safety mechanism is no longer a real concern.
    lob_codes_from_separate_sheet = {} if has_lob_in_master else _load_lob_codes_from_separate_sheet(wb)

    employee_by_emp_no: dict[str, MobitelEmployee] = {e.emp_no: e for e in db.query(MobitelEmployee).all()}
    connection_by_mobile: dict[str, MobitelConnection] = {c.mobile_no: c for c in db.query(MobitelConnection).all()}

    grouped: dict[str, dict] = {}
    pool_rows: list[tuple[str, str]] = []   # (mobile_no, synthetic_emp_no)
    skipped_missing = 0

    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=1000, values_only=True):
        mobile_no = row[mobile_idx] if mobile_idx < len(row) else None
        emp_no = row[emp_no_idx] if emp_no_idx < len(row) else None
        name = row[name_idx] if name_idx < len(row) else None

        if mobile_no is None and emp_no is None and name is None:
            continue

        mobile_no_str = to_str(mobile_no)

        if is_pool_row(emp_no, name):
            if not mobile_no_str:
                skipped_missing += 1
                continue
            pool_rows.append((mobile_no_str, f"POOL-{mobile_no_str}"))
            continue

        emp_no_str = to_str(emp_no)
        raw_name = clean(name)
        base_name, project_label = split_name_and_project_label(raw_name)
        if not mobile_no_str or not emp_no_str or not base_name:
            skipped_missing += 1
            continue

        team_name = clean(row[team_idx]) if team_idx is not None and team_idx < len(row) else None
        lob_code = to_str(row[lob_idx]) if has_lob_in_master and lob_idx < len(row) else lob_codes_from_separate_sheet.get(emp_no_str)

        grouped.setdefault(emp_no_str, {"name": base_name, "team": team_name, "lob_code": lob_code, "mobiles": []})
        # Prefer a row with NO project label for the employee's own name —
        # same reasoning as Dialog Mobile: a project-labeled row's base
        # name is usually identical, but this guards against any stray
        # difference in casing/whitespace across that employee's rows.
        if project_label is None:
            grouped[emp_no_str]["name"] = base_name
        grouped[emp_no_str]["mobiles"].append((mobile_no_str, project_label))

    emp_nos_in_upload = set(grouped.keys()) | {syn for _, syn in pool_rows}
    inserted, updated, revived, retired_employees = 0, 0, 0, 0
    connections_added, connections_retired, connections_reactivated = 0, 0, 0
    conflicts: list[str] = []

    for emp_no, data in grouped.items():
        existing = employee_by_emp_no.get(emp_no)
        if existing is None:
            employee = MobitelEmployee(
                emp_no=emp_no, name=data["name"], lob=data["team"], lob_code=data["lob_code"], is_pool=False,
            )
            db.add(employee)
            db.flush()
            employee_by_emp_no[emp_no] = employee
            inserted += 1
        else:
            employee = existing
            was_deleted = employee.is_deleted
            employee.name = data["name"]
            employee.lob = data["team"]
            if data["lob_code"]:
                employee.lob_code = data["lob_code"]
            employee.is_deleted = False
            employee.is_pool = False
            if was_deleted:
                revived += 1
            else:
                updated += 1

        current_mobiles = {c.mobile_no: c for c in db.query(MobitelConnection).filter(MobitelConnection.employee_id == employee.id).all()}
        new_mobiles = {mobile_no: label for mobile_no, label in data["mobiles"]}

        for mobile_no, project_label in new_mobiles.items():
            conn = connection_by_mobile.get(mobile_no)
            if conn is None:
                db.add(MobitelConnection(
                    employee_id=employee.id, mobile_no=mobile_no,
                    status=MobitelConnectionStatus.active, project_label=project_label,
                ))
                connections_added += 1
            elif conn.employee_id != employee.id:
                conflicts.append(f"{mobile_no}: claimed by a different employee than EMP {emp_no}")
            else:
                if conn.status != MobitelConnectionStatus.active:
                    conn.status = MobitelConnectionStatus.active
                    connections_reactivated += 1
                conn.project_label = project_label

        for mobile_no, conn in current_mobiles.items():
            if mobile_no not in new_mobiles and conn.status == MobitelConnectionStatus.active:
                conn.status = MobitelConnectionStatus.inactive
                connections_retired += 1

    for mobile_no, synthetic_emp_no in pool_rows:
        existing_conn = connection_by_mobile.get(mobile_no)
        if existing_conn is not None and existing_conn.employee_id in {e.id for e in employee_by_emp_no.values() if not e.is_pool}:
            # This number belonged to a REAL employee before and is now
            # showing as unassigned "Pool" — a genuine status change, not
            # something to silently reassign or drop. Confirmed real case:
            # this is exactly how one of the 4 known real reassignment
            # conflicts actually happens (705329362: was Isuru Hasanka's
            # number in June, "NA"/Pool in July).
            owner = next((e for e in employee_by_emp_no.values() if e.id == existing_conn.employee_id), None)
            conflicts.append(
                f"{mobile_no}: file shows this as unassigned 'Pool', but it's currently assigned to "
                f"EMP {owner.emp_no if owner else '?'} ({owner.name if owner else '?'})"
            )
            continue

        existing = employee_by_emp_no.get(synthetic_emp_no)
        if existing is None:
            employee = MobitelEmployee(emp_no=synthetic_emp_no, name="Pool", is_pool=True)
            db.add(employee)
            db.flush()
            employee_by_emp_no[synthetic_emp_no] = employee
            if existing_conn is None:
                db.add(MobitelConnection(employee_id=employee.id, mobile_no=mobile_no))
                connections_added += 1
            inserted += 1
        else:
            was_deleted = existing.is_deleted
            existing.is_deleted = False
            if was_deleted:
                revived += 1

    for emp_no, employee in employee_by_emp_no.items():
        if emp_no not in emp_nos_in_upload and not employee.is_deleted:
            employee.is_deleted = True
            retired_employees += 1

    db.commit()

    return {
        "inserted_employees": inserted,
        "updated_employees": updated,
        "revived_employees": revived,
        "retired_employees": retired_employees,
        "connections_added": connections_added,
        "connections_retired": connections_retired,
        "connections_reactivated": connections_reactivated,
        "skipped_missing_rows": skipped_missing,
        "conflicts": conflicts,
        "lob_code_source": "directly in Master sheet" if has_lob_in_master else (
            "separate LOB sheet" if lob_codes_from_separate_sheet else "unavailable (no LOB sheet found)"
        ),
    }