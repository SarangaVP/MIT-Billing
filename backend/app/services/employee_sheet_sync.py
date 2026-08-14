"""
Syncs Dialog Mobile employees + mobile numbers from an uploaded Excel
sheet, treating that sheet as the FULL, authoritative picture of the
current roster — not just an additive backfill.

Rather than deleting anything (which would violate real constraints —
mobile numbers referenced by past bills, and unique EMP No/mobile-number
constraints that block simply re-inserting under the same identifier
even after a soft delete), this UPDATES existing rows in place:

  - Anyone in the sheet who already exists: their fields are fully
    OVERWRITTEN to match the sheet (not just backfilled when blank).
  - Any of their numbers newly appearing in the sheet: added as active.
  - Any of their numbers NOT in the sheet anymore: marked inactive —
    the row stays, so any past bill referencing it stays valid, but it's
    excluded from future billing.
  - Anyone previously retired (is_deleted) who reappears in the sheet:
    revived automatically.
  - Anyone currently active who is ENTIRELY ABSENT from the sheet: their
    employee record is marked deleted (soft) — no longer in the current
    roster, but nothing about their history is touched.

Shares its row-parsing logic (project label splitting, "General"
line handling, header structure) with import_master_sheet.py —
kept as a single source of truth so the CLI script and the web upload
endpoint can never drift apart.
"""
import uuid

import openpyxl
from sqlalchemy.orm import Session

from app.models.employee import Employee
from app.models.mobile_number import MobileNumber, MobileNumberStatus

ADDITIONAL_BLOCK_MARKER = "Additional common connection"
UNRESOLVED_EMP_NO_PLACEHOLDERS = {"General"}
NO_NUMBER_ASSIGNED_PLACEHOLDERS = {"no", "none", "n/a", "na", "tbc", "pending", "-"}


def is_no_number_placeholder(value) -> bool:
    return isinstance(value, str) and value.strip().lower() in NO_NUMBER_ASSIGNED_PLACEHOLDERS


def split_name_and_project_label(raw_name):
    if raw_name is None:
        return None, None
    if "-" not in raw_name:
        return raw_name.strip(), None
    base, _, label = raw_name.partition("-")
    return base.strip(), label.strip() or None


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.replace("\ufeff", "")
        if value.strip() == "":
            return None
    return value


def to_text(value):
    return None if value is None else str(value)


def load_main_table_rows(xlsx_path: str) -> list[dict]:
    """
    Returns a list of dicts (one per real data row), keyed by normalized
    header name — e.g. row["mobile no"], row["emp no"] — instead of
    fixed positions. Confirmed real risk this replaces: the previous
    version read row[0] through row[8] with ZERO verification that
    column 4 was actually "Cadre", column 7 actually "Email", etc. — if
    the sheet's column order ever changed, this would have silently
    corrupted data with no error at all (the same failure mode already
    confirmed and fixed for the Portal sheet elsewhere in this project).

    Also handles a CONFIRMED real typo in the source file's own header:
    the Email column is literally spelled "Emaill" (double-L), not
    "Email" — both spellings are accepted and mapped to the same field,
    so a future corrected spelling wouldn't silently break this either.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Master sheet"]

    header_row_num = None
    col_map: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if "mobile no" in values and "emp no" in values and "name" in values:
            header_row_num = row[0].row
            col_map = {v: i for i, v in enumerate(values) if v}
            break
    if header_row_num is None:
        raise ValueError("Could not find a header row containing 'Mobile No', 'EMP No', and 'Name' in the first 5 rows")

    email_idx = col_map.get("email", col_map.get("emaill"))   # confirmed real header typo: "Emaill"

    field_columns = {
        "mobile no": col_map.get("mobile no"),
        "emp no": col_map.get("emp no"),
        "name": col_map.get("name"),
        "lob": col_map.get("lob"),
        "cadre": col_map.get("cadre"),
        "credit limit": col_map.get("credit limit"),
        "level": col_map.get("level"),
        "email": email_idx,
        "resignation": col_map.get("resignation"),
    }

    all_rows = list(ws.iter_rows(min_row=header_row_num + 1, max_row=ws.max_row, values_only=True))

    marker_idx = None
    for i, row in enumerate(all_rows):
        for cell in row:
            if isinstance(cell, str) and cell.strip() == ADDITIONAL_BLOCK_MARKER:
                marker_idx = i
                break
        if marker_idx is not None:
            break

    main_rows = all_rows[:marker_idx] if marker_idx is not None else all_rows

    def row_to_dict(row: tuple) -> dict:
        return {
            field: (row[idx] if idx is not None and idx < len(row) else None)
            for field, idx in field_columns.items()
        }

    return [row_to_dict(row) for row in main_rows]


def sync_employee_sheet(db: Session, xlsx_path: str) -> dict:
    """
    Returns a structured result dict — used by both the CLI script and the
    web upload endpoint, so their reported summaries are always identical.
    """
    main_rows = load_main_table_rows(xlsx_path)

    employee_by_emp_no: dict[str, Employee] = {e.emp_no: e for e in db.query(Employee).all()}
    number_by_mobile: dict[str, MobileNumber] = {n.mobile_no: n for n in db.query(MobileNumber).all()}

    # Group real (non-"General") rows by EMP No first, same as before —
    # one employee can have multiple rows (multiple numbers).
    grouped: dict[str, dict] = {}
    general_line_rows: list[tuple[str, str, str | None]] = []   # (mobile_no, base_name, source_row)
    skipped_missing = 0
    conflicts: list[tuple[str, str, str]] = []

    for row in main_rows:
        raw_mobile_no = clean(row["mobile no"])
        emp_no = to_text(clean(row["emp no"]))
        raw_name = clean(row["name"])
        base_name, project_label = split_name_and_project_label(raw_name)

        if not emp_no:
            skipped_missing += 1
            continue

        if emp_no in UNRESOLVED_EMP_NO_PLACEHOLDERS:
            mobile_no_str = to_text(raw_mobile_no)
            if not mobile_no_str:
                skipped_missing += 1
                continue
            general_line_rows.append((mobile_no_str, base_name, row))
            continue

        grouped.setdefault(emp_no, {"rows": [], "first": row, "clean_name_row": None})
        if project_label is None and grouped[emp_no]["clean_name_row"] is None:
            grouped[emp_no]["clean_name_row"] = row

        if raw_mobile_no is None:
            skipped_missing += 1
            continue
        if is_no_number_placeholder(raw_mobile_no):
            continue

        grouped[emp_no]["rows"].append((to_text(raw_mobile_no), project_label))

    emp_nos_in_upload: set[str] = set(grouped.keys())
    inserted, updated, revived, retired_employees = 0, 0, 0, 0
    numbers_added, numbers_retired, numbers_reactivated = 0, 0, 0

    def upsert_employee_fields(employee: Employee, source_row, synthetic_name: str | None = None, is_general: bool = False):
        clean_name, _ = split_name_and_project_label(clean(source_row["name"]))
        employee.name = synthetic_name if synthetic_name is not None else clean_name
        employee.lob = clean(source_row["lob"])
        employee.cadre = clean(source_row["cadre"])
        employee.credit_limit = clean(source_row["credit limit"])
        employee.level = clean(source_row["level"])
        employee.email = clean(source_row["email"])
        employee.resignation = clean(source_row["resignation"])
        employee.is_general_line = is_general

    # ---- Regular employees ----
    for emp_no, data in grouped.items():
        source_row = data["clean_name_row"] or data["first"]
        new_numbers = {mobile_no: label for mobile_no, label in data["rows"]}

        existing = employee_by_emp_no.get(emp_no)
        if existing is None:
            employee = Employee(emp_no=emp_no)
            upsert_employee_fields(employee, source_row)
            db.add(employee)
            db.flush()
            employee_by_emp_no[emp_no] = employee
            inserted += 1
        else:
            employee = existing
            was_deleted = employee.is_deleted
            upsert_employee_fields(employee, source_row)
            employee.is_deleted = False
            if was_deleted:
                revived += 1
            else:
                updated += 1

        current_numbers = {n.mobile_no: n for n in employee.mobile_numbers}

        for mobile_no, project_label in new_numbers.items():
            existing_number = number_by_mobile.get(mobile_no)
            if existing_number is None:
                db.add(MobileNumber(
                    employee_id=employee.id, mobile_no=mobile_no,
                    is_primary=not current_numbers, project_label=project_label,
                    status=MobileNumberStatus.active,
                ))
                numbers_added += 1
            elif existing_number.employee_id != employee.id:
                conflicts.append((emp_no, mobile_no, f"claimed by a different employee"))
            else:
                if existing_number.status != MobileNumberStatus.active:
                    existing_number.status = MobileNumberStatus.active
                    numbers_reactivated += 1
                existing_number.project_label = project_label

        for mobile_no, number in current_numbers.items():
            if mobile_no not in new_numbers and number.status == MobileNumberStatus.active:
                number.status = MobileNumberStatus.inactive
                numbers_retired += 1

    # ---- "General" lines — each is its own synthetic employee ----
    for mobile_no, base_name, source_row in general_line_rows:
        synthetic_emp_no = f"GENERAL-{mobile_no}"
        emp_nos_in_upload.add(synthetic_emp_no)
        existing = employee_by_emp_no.get(synthetic_emp_no)

        if existing is None:
            employee = Employee(emp_no=synthetic_emp_no)
            upsert_employee_fields(employee, source_row, synthetic_name=base_name, is_general=True)
            db.add(employee)
            db.flush()
            employee_by_emp_no[synthetic_emp_no] = employee
            existing_number = number_by_mobile.get(mobile_no)
            if existing_number is None:
                db.add(MobileNumber(employee_id=employee.id, mobile_no=mobile_no, is_primary=True))
                numbers_added += 1
            inserted += 1
        else:
            was_deleted = existing.is_deleted
            upsert_employee_fields(existing, source_row, synthetic_name=base_name, is_general=True)
            existing.is_deleted = False
            if was_deleted:
                revived += 1
            else:
                updated += 1

    # ---- Anyone active in the DB but entirely absent from this upload ----
    for emp_no, employee in employee_by_emp_no.items():
        if emp_no not in emp_nos_in_upload and not employee.is_deleted:
            employee.is_deleted = True
            retired_employees += 1

    db.commit()

    return {
        "inserted_employees": inserted,
        "updated_employees": updated,
        "revived_employees": revived,
        "retired_employees": retired_employees,
        "numbers_added": numbers_added,
        "numbers_retired": numbers_retired,
        "numbers_reactivated": numbers_reactivated,
        "skipped_missing_rows": skipped_missing,
        "conflicts": conflicts,
    }