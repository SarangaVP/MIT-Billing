"""
Parses the 'Summary' sheet (and, if needed, 'Team cost' for LOB codes)
from the SLT team package Excel — uploaded FRESH every month alongside
the PDF, not seeded once. This IS the source of truth for that month's
employee/package/team allocation; there's no persistent employee roster.

CONFIRMED two real file formats seen in practice:
  - Older format: Summary sheet has Name, Package, Team, Amount only.
    LOB code must be cross-referenced from the separate 'Team cost'
    sheet, by team name.
  - Newer format: Summary sheet has a 5th column, LOB, with the numeric
    code directly on each row — no need to cross-reference the other
    sheet at all. Verified the two formats' LOB codes agree exactly.

Reads columns by HEADER NAME, not fixed position — confirmed this
file's real header row is at row 2, not row 1 (row 1 is blank), which a
fixed "skip row 1" assumption would have silently misread as data.

Footer/summary rows (Cess, SSCL, a Package-count pivot table) are
skipped by name-marker, not by row position, since their exact row
count can shift between files.
"""
FOOTER_MARKERS = {"cess", "sscl", "package", "total", "work & learn 100gb", "20gb anytime data"}


def clean_text(value) -> str | None:
    if value is None:
        return None
    cleaned = str(value).replace("\ufeff", "").strip()
    cleaned = " ".join(cleaned.split())  # collapse internal double-spaces
    return cleaned or None


def _find_header_row(ws, required_headers: list[str]) -> int:
    for row in ws.iter_rows(min_row=1, max_row=10):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if values[: len(required_headers)] == required_headers:
            return row[0].row
    raise ValueError(f"Could not find header row {required_headers} in sheet '{ws.title}'")


def _load_lob_codes_from_team_cost(wb) -> dict[str, str]:
    """Fallback for the OLDER format — cross-references team name -> LOB code."""
    if "Team cost" not in wb.sheetnames:
        return {}
    ws = wb["Team cost"]
    header_row = _find_header_row(ws, ["team", "sum of amount", "lob"])
    codes: dict[str, str] = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        team, lob_code = row[0], row[2] if len(row) > 2 else None
        team_clean = clean_text(team)
        if team_clean and lob_code is not None and team_clean.lower() not in {"cess", "sscl", "grand total"}:
            codes[team_clean] = str(lob_code)
    return codes


def parse_summary_excel(xlsx_path: str) -> list[dict]:
    """
    Returns a list of {name, team, lob_code, package_name, package_price}
    for every real employee row in the 'Summary' sheet — this is the
    complete, fresh allocation for THIS specific month's bill.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Summary"]
    header_row = _find_header_row(ws, ["name", "package", "team", "amount"])

    header_values = [str(c.value).strip().lower() if c.value else None for c in ws[header_row]]
    has_lob_column = "lob" in header_values
    lob_col_idx = header_values.index("lob") if has_lob_column else None

    lob_codes_by_team = {} if has_lob_column else _load_lob_codes_from_team_cost(wb)

    results = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name, package, team, amount = row[0], row[1], row[2], row[3]

        name_clean = clean_text(name)
        if name_clean is None:
            continue

        if name_clean.lower() in FOOTER_MARKERS:
            continue

        package_clean, team_clean = clean_text(package), clean_text(team)
        if not package_clean or not team_clean or amount is None:
            continue

        if has_lob_column and lob_col_idx is not None and lob_col_idx < len(row) and row[lob_col_idx] is not None:
            lob_code = str(row[lob_col_idx])
        else:
            lob_code = lob_codes_by_team.get(team_clean)

        results.append({
            "name": name_clean,
            "team": team_clean,
            "lob_code": lob_code,
            "package_name": package_clean,
            "package_price": float(amount),
        })

    return results