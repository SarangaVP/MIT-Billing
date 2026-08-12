"""
Parses the "Bill" sheet from the Dialog Data Bucket Excel export — per-
connection usage detail (Allocation/Usage/Remaining GB, Pay Go status).
Mirrors Mobitel's Portal parser. This is a monthly SNAPSHOT, matched to a
bill period's line items by mobile number — not written into employees.

Reads columns by HEADER NAME and detects the header row dynamically,
rather than assuming it's always row 1 with a fixed 8-column layout —
same defensive pattern applied across every sheet in this project after
confirming real header/column shifts happen in this workbook family
(the Portal sheet in the Mobitel workbook, and Dialog Mobile's own
Master sheet, both had real examples of this).

NOTE: Allocation/Usage/Remaining GB are stored as raw strings, not
numbers — real data includes non-numeric values like "UNLIMITED" and
"NaN" (confirmed in the actual file), which would break a numeric column.
"""
import openpyxl

REQUIRED_HEADERS = ["mobile number", "name"]


def _mobile_no_to_text(value) -> str:
    if isinstance(value, float):
        return str(int(value))
    return str(value)


def _find_header_row_and_columns(ws) -> tuple[int, dict[str, int]]:
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if values[: len(REQUIRED_HEADERS)] == REQUIRED_HEADERS:
            col_map = {v: i for i, v in enumerate(values) if v}
            return row[0].row, col_map
    raise ValueError("Could not find the Bill sheet's header row (expected 'Mobile Number', 'Name')")


def parse_bill_sheet(xlsx_path: str) -> dict:
    """Returns {mobile_no: {field: value, ...}}."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Bill"]

    header_row, col_map = _find_header_row_and_columns(ws)

    field_columns = {
        "mobile_no": col_map.get("mobile number"),
        "allocation_gb": col_map.get("allocation gb"),
        "usage_gb": col_map.get("usage gb"),
        "remaining_gb": col_map.get("remaining gb"),
        "pay_go_status": col_map.get("pay go status"),
        "bill_cycle": col_map.get("bill cycle"),
    }

    if field_columns["mobile_no"] is None:
        raise ValueError("Could not find a 'Mobile Number' column in the Bill sheet's header row")

    result = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        mobile_idx = field_columns["mobile_no"]
        if mobile_idx >= len(row) or row[mobile_idx] is None:
            continue

        def get(field):
            idx = field_columns[field]
            return row[idx] if idx is not None and idx < len(row) else None

        allocation_gb, usage_gb, remaining_gb = get("allocation_gb"), get("usage_gb"), get("remaining_gb")
        result[_mobile_no_to_text(row[mobile_idx])] = {
            "allocation_gb": str(allocation_gb) if allocation_gb is not None else None,
            "usage_gb": str(usage_gb) if usage_gb is not None else None,
            "remaining_gb": str(remaining_gb) if remaining_gb is not None else None,
            "pay_go_status": get("pay_go_status"),
            "bill_cycle": get("bill_cycle"),
        }

    return result