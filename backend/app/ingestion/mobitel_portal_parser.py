"""
Parses the "Portal" sheet from the Mobitel data bucket Excel export — a raw
per-SIM technical export (IMSI, allocated/available/utilized data volume,
daily limits, member status). This is separate from the "Summary" sheet
(used for employee seeding) — Portal has no EMP No/LOB at all, only the
technical usage detail.

Reads columns by HEADER NAME, not fixed position, and detects the header
row dynamically (searching for "IMSI Number") — confirmed real file
format differences between two actual exports:
  - Older format: a title row above the header, header at row 2, Mobile
    Number at position 3.
  - Newer format: NO title row (header directly at row 1), an extra
    unlabeled column inserted before Mobile Number, shifting it to
    position 4.
A fixed "row 3 = data, index 3 = mobile_no" assumption would have
silently skipped the first real row AND matched every row's usage data
to the wrong mobile number against the newer file — confirmed by
checking the real header text and data side-by-side.

This is a monthly SNAPSHOT, not persistent employee data — matched to a
bill period's line items by mobile number, not written into mobitel_employees.
"""
import openpyxl

REQUIRED_HEADERS = ["imsi number", "iccid", "name"]


def _mobile_no_to_text(value) -> str:
    if isinstance(value, float):
        return str(int(value))
    return str(value)


def _find_header_row_and_columns(ws) -> tuple[int, dict[str, int]]:
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [str(c.value).strip().lower() if c.value else None for c in row]
        if values[: len(REQUIRED_HEADERS)] == REQUIRED_HEADERS:
            col_map = {v: i for i, v in enumerate(values) if v}
            return row[0].row, col_map
    raise ValueError("Could not find the Portal sheet's header row (expected 'IMSI Number', 'ICCID', 'Name')")


def parse_portal_sheet(xlsx_path: str) -> dict:
    """Returns {mobile_no: {field: value, ...}}, keyed for easy lookup at import time."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Portal"]

    header_row, col_map = _find_header_row_and_columns(ws)

    field_columns = {
        "imsi_number": col_map.get("imsi number"),
        "mobile_no": col_map.get("mobile number"),
        "data_volume_mb": col_map.get("data volume (mb)"),
        "available_data_volume_mb": col_map.get("available data volume (mb)"),
        "utilized_data_volume_mb": col_map.get("utilized data volume (mb)"),
        "daily_limit_mb": col_map.get("daily limit(mb)"),
        "utilized_daily_limit_mb": col_map.get("utilized daily limit (mb)"),
        "member_status": col_map.get("member status"),
        "top_up_mb": col_map.get("top up (mb)"),
        "utilized_topup_mb": col_map.get("utilized units topup (mb)"),
    }

    if field_columns["mobile_no"] is None:
        raise ValueError("Could not find a 'Mobile Number' column in the Portal sheet's header row")

    result = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        mobile_idx = field_columns["mobile_no"]
        if mobile_idx >= len(row) or row[mobile_idx] is None:
            continue

        mobile_no = _mobile_no_to_text(row[mobile_idx])

        def get(field):
            idx = field_columns[field]
            return row[idx] if idx is not None and idx < len(row) else None

        imsi = get("imsi_number")
        result[mobile_no] = {
            "imsi_number": str(imsi) if imsi is not None else None,
            "data_volume_mb": get("data_volume_mb"),
            "available_data_volume_mb": get("available_data_volume_mb"),
            "utilized_data_volume_mb": get("utilized_data_volume_mb"),
            "daily_limit_mb": get("daily_limit_mb"),
            "utilized_daily_limit_mb": get("utilized_daily_limit_mb"),
            "member_status": get("member_status"),
            "top_up_mb": get("top_up_mb"),
            "utilized_topup_mb": get("utilized_topup_mb"),
        }

    return result